from openpyxl import Workbook, load_workbook
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Side, Border


class GenerateExcel: 
    def __init__(self) -> None:
        # Create work book 
        self.wb = Workbook()
        self.ws = self.wb.active 
    
        # Set title
        self.ws.title = "Products"

        # Create border for cells 
        thin = Side(border_style="thin", color="000000")
        self.border = Border(top=thin, right=thin, bottom=thin, left=thin)

    def add_headings(self, headings: list[str]) -> None:
        self.ws.append(headings) 

        # Add styles
        for col in range(1, len(headings) + 1):
            self.ws.column_dimensions[get_column_letter(col)].width = 30 
            self.ws[get_column_letter(col) + "1"].font = Font(bold=True, color="00FFFFFF") 
            self.ws[get_column_letter(col) + "1"].border = self.border 
            self.ws[get_column_letter(col) + "1"].fill = PatternFill(start_color="0057bc5f", end_color="0057bc5f", fill_type="solid")
    
    def add_data(self, len_heading: int, data: list[dict[str, str | int]]) -> None: 
        for row in data: 
            self.ws.append(list(row.values()))

        # Add styles 
        for col in range(1, len_heading + 1): 
            for row in range(2, len(data) + 2): 
                self.ws[get_column_letter(col) + str(row)].border = self.border 

    def generate_file(self, filename: str, headings: list[str], data: list[dict[str, str | int]]): 
        self.add_headings(headings)
        self.add_data(len(headings), data)

        self.wb.save(f"{filename}.xlsx")

    def generate_file_buffer(self, headings: list[str], data: list[dict[str, str | int]]): 
        self.add_headings(headings) 
        self.add_data(len(headings), data)

        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0) 
        return buffer


class ReadExcel: 
    def __init__(self, file: bytes): 
        self.wb = load_workbook(BytesIO(file))
        self.ws = self.wb.active 

    def read_file(self):
        excel_data = [] 
        for row in range(2, self.ws.max_row + 1):
            cells = []
            for col in range(1, self.ws.max_column + 1):
                cells.append(get_column_letter(col) + str(row))
                
            data = {
                "title": self.ws[cells[0]].value, 
                "category": self.ws[cells[1]].value, 
                "price": self.ws[cells[2]].value, 
                "image": self.ws[cells[3]].value
            }
            excel_data.append(data)

        return excel_data
    